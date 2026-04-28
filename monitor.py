#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
╔══════════════════════════════════════════════════════════╗
║       ICETEX Social Media Monitor  v2.0                  ║
║       Facebook · X (Twitter) · Instagram                 ║
╠══════════════════════════════════════════════════════════╣
║  Analiza:                                                ║
║   • Comentarios de usuarios en publicaciones             ║
║   • Si ICETEX respondió o no                             ║
║   • Tiempo de respuesta vs SLA por plataforma            ║
║   • Calidad/pertinencia de la respuesta (mejorada)       ║
║   • 9 categorías de temas                                ║
║  Genera:                                                 ║
║   • Reporte HTML con gráficas interactivas               ║
║   • Exportación a Excel (.xlsx)                          ║
║   • Alertas automáticas por correo electrónico           ║
╚══════════════════════════════════════════════════════════╝
"""

import json
import os
import sys
import time
import html as html_lib
import smtplib

# Forzar UTF-8 en la consola de Windows para que los emojis no causen error
if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
        sys.stderr.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from datetime import datetime, timezone, timedelta
from typing import Optional, List, Dict
import requests

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_OK = True
except ImportError:
    EXCEL_OK = False

# ──────────────────────────────────────────────────────────
# CONFIGURACIÓN
# ──────────────────────────────────────────────────────────
CONFIG_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "config.json")

DEFAULT_CONFIG = {
    "twitter":   {"bearer_token": "TU_BEARER_TOKEN_AQUI",    "username": "ICETEXColombia", "sla_horas": 4},
    "facebook":  {"page_id": "ICETEXColombia",                "access_token": "TU_PAGE_ACCESS_TOKEN_AQUI", "sla_horas": 8},
    "instagram": {"user_id": "", "access_token": "",          "username": "icetex_colombia", "sla_horas": 12},
    "settings":  {"days_back": 7, "max_posts_per_platform": 10},
    "alertas":   {
        "activo": False,
        "smtp_host": "smtp.gmail.com", "smtp_port": 587,
        "usuario": "tu_correo@gmail.com", "password": "tu_contraseña_de_aplicacion",
        "destinatarios": ["destinatario@icetex.gov.co"],
        "umbral_sin_respuesta": 5
    }
}

# ──────────────────────────────────────────────────────────
# TEMAS (9 categorías)
# ──────────────────────────────────────────────────────────
TEMAS = {
    "crédito / pago": [
        "crédito", "credito", "préstamo", "prestamo", "cuota", "pago",
        "deuda", "saldo", "financiamiento", "interés", "interes",
        "refinanciación", "refinanciacion", "abono", "mora", "cartera",
        "cobro", "factura", "extracto"
    ],
    "beca / condonación": [
        "beca", "condonación", "condonacion", "subsidio", "beneficio",
        "ser pilo", "generación e", "generacion e", "spa", "gratuidad",
        "matrícula", "matricula", "exención", "exencion"
    ],
    "desembolso / giro": [
        "desembolso", "giro", "transferencia", "consignación", "consignacion",
        "cuenta bancaria", "banco", "nequi", "daviplata", "pse",
        "no llega", "no han girado", "cuándo giran", "cuando giran"
    ],
    "documentos / trámites": [
        "documento", "requisito", "certificado", "paz y salvo", "carta",
        "constancia", "formulario", "trámite", "tramite", "solicitud",
        "registro", "acta", "diploma", "cédula", "cedula", "rut", "comprobante"
    ],
    "calificación / renovación": [
        "calificación", "calificacion", "renovar", "renovación", "renovacion",
        "puntaje", "nota", "promedio", "icfes", "académico", "academico",
        "semestre", "periodo", "corte", "período"
    ],
    "sistema / plataforma": [
        "sistema", "plataforma", "no carga", "error", "no funciona",
        "página", "pagina", "app", "aplicación", "aplicacion", "portal",
        "caído", "caido", "lento", "colgado", "falla", "contraseña",
        "login", "acceso", "ingresar", "registrar"
    ],
    "queja / reclamo": [
        "queja", "reclamo", "denuncia", "inconformidad", "insatisfecho",
        "malestar", "pésimo", "pesimo", "terrible", "mal servicio",
        "abuso", "negligencia", "vergüenza", "verguenza", "inaceptable",
        "indignante", "estafa", "perjuicio"
    ],
    "urgencia": [
        "urgente", "urgencia", "inmediato", "pronto", "cuándo", "cuando",
        "demora", "retraso", "tardanza", "hace días", "hace meses", "llevo",
        "esperando", "sin respuesta", "ignorando", "nadie responde",
        "vencimiento", "vence", "plazo", "fecha límite", "fecha limite"
    ],
    "información general": [
        "información", "informacion", "cómo", "como", "dónde", "donde",
        "requisitos", "proceso", "pasos", "ayuda", "explica", "orienta",
        "guía", "guia", "indicar", "saber", "consultar"
    ]
}

EMOCIONES = {
    "Crítico": [
        "queja", "reclamo", "denuncia", "estafa", "abuso", "escándalo", "escandalo",
        "inaceptable", "indignante", "vergüenza", "verguenza", "perjuicio",
        "negligencia", "corrupción", "corrupcion", "demanda", "tutela",
        "fraude", "robo", "mentira", "engaño", "engano", "incumplimiento"
    ],
    "Negativo": [
        "malo", "mal ", "terrible", "pésimo", "pesimo", "horrible", "fatal",
        "molesto", "frustrado", "decepcionado", "decepción", "decepcion",
        "insatisfecho", "inconformidad", "malestar", "no funciona",
        "no responden", "nadie responde", "sin respuesta",
        "demora", "retraso", "tardanza", "cansado", "harto"
    ],
    "Positivo": [
        "gracias", "excelente", "genial", "perfecto", "bien ", "bueno", "buena",
        "feliz", "contento", "satisfecho", "encantado", "maravilloso",
        "increíble", "increible", "fantástico", "fantastico", "agradecido",
        "agradezco", "aprecio", "resolvieron", "solucionaron",
        "atendieron", "eficiente", "buen servicio"
    ],
}

FRASES_GENERICAS = [
    "gracias por contactarnos", "gracias por escribirnos", "gracias por comunicarte",
    "te invitamos a ingresar", "visita nuestra página", "ingresa a",
    "por favor escríbenos", "escríbenos al correo", "comunícate al",
    "lamentamos el inconveniente", "entendemos tu situación",
    "con gusto te ayudamos", "estamos para servirte", "quedamos atentos",
    "bienvenido", "hola! gracias", "hola, gracias", "buenos días, gracias",
    "buenas tardes, gracias", "buenas noches, gracias", "con mucho gusto",
    "en breve", "a la mayor brevedad", "nos pondremos en contacto"
]

STOPWORDS = {
    "el", "la", "los", "las", "un", "una", "y", "o", "de", "en", "que",
    "a", "por", "con", "su", "al", "del", "es", "se", "me", "te", "nos",
    "hay", "para", "pero", "como", "más", "mas", "muy", "ya", "si", "no"
}

# ──────────────────────────────────────────────────────────
# UTILIDADES
# ──────────────────────────────────────────────────────────
def cargar_config() -> Dict:
    if not os.path.exists(CONFIG_FILE):
        with open(CONFIG_FILE, "w", encoding="utf-8") as f:
            json.dump(DEFAULT_CONFIG, f, indent=2, ensure_ascii=False)
        print(f"\n✅ Archivo config.json creado en:\n   {CONFIG_FILE}")
        print("⚠️  Edita ese archivo con tus credenciales antes de ejecutar de nuevo.\n")
        sys.exit(0)
    with open(CONFIG_FILE, encoding="utf-8") as f:
        return json.load(f)


def parsear_fecha(s: str) -> Optional[datetime]:
    for fmt in [
        "%Y-%m-%dT%H:%M:%S.%fZ", "%Y-%m-%dT%H:%M:%SZ",
        "%Y-%m-%dT%H:%M:%S+0000", "%Y-%m-%dT%H:%M:%S%z"
    ]:
        try:
            dt = datetime.strptime(s, fmt)
            return dt if dt.tzinfo else dt.replace(tzinfo=timezone.utc)
        except ValueError:
            continue
    return None


def fmt_tiempo(minutos: Optional[float]) -> str:
    if minutos is None:
        return "—"
    if minutos < 60:
        return f"{int(minutos)} min"
    if minutos < 1440:
        return f"{minutos/60:.1f} h"
    return f"{minutos/1440:.1f} días"


def clasificar_tema(texto: str) -> str:
    t = texto.lower()
    mejor_tema, mejor_score = "otro", 0
    for tema, palabras in TEMAS.items():
        score = sum(1 for p in palabras if p in t)
        if score > mejor_score:
            mejor_score = score
            mejor_tema = tema
    return mejor_tema


def detectar_emocion(texto: str) -> str:
    t = texto.lower()
    for emocion in ["Crítico", "Negativo", "Positivo"]:
        if any(p in t for p in EMOCIONES[emocion]):
            return emocion
    return "Neutro"


def evaluar_calidad(comentario: str, respuesta: str) -> Dict:
    """
    Score: 0=sin respuesta  1=genérica  2=parcial  3=específica
    Considera: frases genéricas, coincidencia de tema, palabras del comentario
    en la respuesta, longitud.
    """
    if not respuesta or not respuesta.strip():
        return {"score": 0, "nivel": "Sin respuesta", "motivo": "ICETEX no respondió"}

    r = respuesta.lower()
    genericas = sum(1 for f in FRASES_GENERICAS if f in r)
    tema = clasificar_tema(comentario)
    palabras_tema = TEMAS.get(tema, [])
    coincidencias_tema = sum(1 for p in palabras_tema if p in r)
    longitud = len(respuesta.strip())

    palabras_com = {w for w in comentario.lower().split() if len(w) > 3 and w not in STOPWORDS}
    coincidencias_directas = sum(1 for p in palabras_com if p in r)

    if genericas >= 2 and coincidencias_tema == 0 and coincidencias_directas == 0:
        return {"score": 1, "nivel": "Genérica", "motivo": "Respuesta automática sin abordar el tema"}
    if genericas >= 1 and longitud < 120 and coincidencias_directas < 2:
        return {"score": 1, "nivel": "Genérica", "motivo": "Respuesta corta y genérica"}
    if (coincidencias_tema > 0 or coincidencias_directas >= 3) and longitud > 150:
        return {"score": 3, "nivel": "Específica", "motivo": "Respuesta aborda el tema del usuario"}
    if coincidencias_directas >= 2 or (longitud > 200 and genericas == 0):
        return {"score": 3, "nivel": "Específica", "motivo": "Respuesta personalizada"}
    return {"score": 2, "nivel": "Parcial", "motivo": "Respuesta presente pero poco específica"}


# ──────────────────────────────────────────────────────────
# MÓDULO X (TWITTER)
# ──────────────────────────────────────────────────────────
class TwitterMonitor:
    BASE = "https://api.twitter.com/2"

    def __init__(self, bearer_token: str, username: str):
        self.headers = {"Authorization": f"Bearer {bearer_token}"}
        self.username = username.lstrip("@")
        self.user_id: Optional[str] = None

    def _get(self, path: str, params: Dict = None) -> Optional[Dict]:
        try:
            r = requests.get(f"{self.BASE}{path}", headers=self.headers, params=params, timeout=15)
            if r.status_code == 429:
                print("  ⚠️  Rate limit X. Esperando 60 s...")
                time.sleep(60)
                r = requests.get(f"{self.BASE}{path}", headers=self.headers, params=params, timeout=15)
            if r.status_code != 200:
                print(f"  ✗ X API {r.status_code}: {r.text[:200]}")
                return None
            return r.json()
        except Exception as e:
            print(f"  ✗ Conexión X: {e}")
            return None

    def get_user_id(self) -> Optional[str]:
        d = self._get(f"/users/by/username/{self.username}")
        if d and "data" in d:
            self.user_id = d["data"]["id"]
            return self.user_id
        return None

    def get_tweets(self, max_results: int, days_back: int) -> List[Dict]:
        if not self.user_id and not self.get_user_id():
            return []
        start = (datetime.now(timezone.utc) - timedelta(days=days_back)).strftime("%Y-%m-%dT%H:%M:%SZ")
        d = self._get(f"/users/{self.user_id}/tweets", {
            "max_results": min(max_results, 100),
            "start_time": start,
            "tweet.fields": "created_at,public_metrics,conversation_id,text",
            "exclude": "retweets,replies"
        })
        return d.get("data", []) if d else []

    def get_replies(self, conversation_id: str) -> List[Dict]:
        d = self._get("/tweets/search/recent", {
            "query": f"conversation_id:{conversation_id} -is:retweet",
            "max_results": 100,
            "tweet.fields": "created_at,author_id,text,in_reply_to_user_id",
            "expansions": "author_id",
            "user.fields": "username"
        })
        if not d or "data" not in d:
            return []
        users = {u["id"]: u["username"] for u in d.get("includes", {}).get("users", [])}
        for t in d["data"]:
            t["username"] = users.get(t["author_id"], "desconocido")
        return d["data"]

    def analizar(self, max_posts: int, days_back: int, sla_horas: float) -> List[Dict]:
        print(f"\n🐦 Analizando X (@{self.username}) — SLA: {sla_horas}h...")
        tweets = self.get_tweets(max_posts, days_back)
        if not tweets:
            print("  Sin tweets recientes o token inválido.")
            return []

        resultados = []
        for tw in tweets:
            print(f"  → Tweet {tw['id'][:12]}... ", end="", flush=True)
            replies = self.get_replies(tw["conversation_id"])
            time.sleep(1)

            icetex_r = [r for r in replies if r["username"].lower() == self.username.lower()]
            user_r   = [r for r in replies if r["username"].lower() != self.username.lower()]

            comentarios = []
            for com in user_r:
                respondido = False; t_resp_min = None; txt_resp = None
                for resp in icetex_r:
                    if resp.get("in_reply_to_user_id") == com["author_id"]:
                        respondido = True
                        tc = parsear_fecha(com["created_at"])
                        tr = parsear_fecha(resp["created_at"])
                        if tc and tr:
                            t_resp_min = (tr - tc).total_seconds() / 60
                        txt_resp = resp["text"]
                        break
                comentarios.append({
                    "usuario": f"@{com['username']}",
                    "texto": com["text"],
                    "fecha": com["created_at"],
                    "respondido": respondido,
                    "tiempo_respuesta_min": t_resp_min,
                    "texto_respuesta": txt_resp,
                    "tema": clasificar_tema(com["text"]),
                    "calidad": evaluar_calidad(com["text"], txt_resp or ""),
                    "emocion": detectar_emocion(com["text"]),
                    "cumple_sla": (t_resp_min <= sla_horas * 60) if t_resp_min is not None else None
                })

            m = tw.get("public_metrics", {})
            resultados.append({
                "id": tw["id"], "texto": tw["text"], "fecha": tw["created_at"],
                "likes": m.get("like_count", 0), "shares": m.get("retweet_count", 0),
                "replies_total": m.get("reply_count", 0), "comentarios": comentarios,
                "url": f"https://x.com/{self.username}/status/{tw['id']}"
            })
            print(f"{len(comentarios)} comentarios de usuarios")
        return resultados


# ──────────────────────────────────────────────────────────
# MÓDULO FACEBOOK
# ──────────────────────────────────────────────────────────
class FacebookMonitor:
    BASE = "https://graph.facebook.com/v19.0"

    def __init__(self, page_id: str, access_token: str):
        self.page_id = page_id
        self.token = access_token
        self.numeric_page_id: Optional[str] = None

    def _get(self, endpoint: str, params: Dict = None) -> Optional[Dict]:
        params = dict(params or {})
        params["access_token"] = self.token
        try:
            r = requests.get(f"{self.BASE}/{endpoint}", params=params, timeout=15)
            if r.status_code != 200:
                print(f"  ✗ FB API {r.status_code}: {r.text[:300]}")
                return None
            return r.json()
        except Exception as e:
            print(f"  ✗ Conexión FB: {e}")
            return None

    def get_page_id(self) -> Optional[str]:
        d = self._get(self.page_id, {"fields": "id,name"})
        if d and "id" in d:
            self.numeric_page_id = d["id"]
            print(f"  ✓ Página: {d.get('name')} (ID {self.numeric_page_id})")
            return self.numeric_page_id
        return None

    def get_posts(self, limit: int, days_back: int) -> List[Dict]:
        since = int((datetime.now() - timedelta(days=days_back)).timestamp())
        d = self._get(f"{self.page_id}/posts", {
            "fields": "id,message,created_time,likes.limit(1).summary(true),shares,comments.limit(1).summary(true)",
            "limit": limit, "since": since
        })
        return d.get("data", []) if d else []

    def get_comments(self, post_id: str) -> List[Dict]:
        d = self._get(f"{post_id}/comments", {
            "fields": "id,from,message,created_time,comments{id,from,message,created_time}",
            "limit": 100, "filter": "stream"
        })
        return d.get("data", []) if d else []

    def analizar(self, max_posts: int, days_back: int, sla_horas: float) -> List[Dict]:
        print(f"\n📘 Analizando Facebook ({self.page_id}) — SLA: {sla_horas}h...")
        if not self.get_page_id():
            print("  Sin acceso o token inválido.")
            return []

        posts = self.get_posts(max_posts, days_back)
        if not posts:
            print("  Sin publicaciones recientes.")
            return []

        resultados = []
        for post in posts:
            print(f"  → Post {post['id'][-10:]}... ", end="", flush=True)
            raw_comments = self.get_comments(post["id"])

            comentarios = []
            for com in raw_comments:
                if com.get("from", {}).get("id") == self.numeric_page_id:
                    continue
                sub = com.get("comments", {}).get("data", [])
                respondido = False; t_resp_min = None; txt_resp = None
                for s in sub:
                    if s.get("from", {}).get("id") == self.numeric_page_id:
                        respondido = True
                        tc = parsear_fecha(com["created_time"])
                        tr = parsear_fecha(s["created_time"])
                        if tc and tr:
                            t_resp_min = (tr - tc).total_seconds() / 60
                        txt_resp = s.get("message", "")
                        break
                texto_com = com.get("message", "")
                comentarios.append({
                    "usuario": com.get("from", {}).get("name", "Usuario"),
                    "texto": texto_com,
                    "fecha": com.get("created_time", ""),
                    "respondido": respondido,
                    "tiempo_respuesta_min": t_resp_min,
                    "texto_respuesta": txt_resp,
                    "tema": clasificar_tema(texto_com),
                    "calidad": evaluar_calidad(texto_com, txt_resp or ""),
                    "emocion": detectar_emocion(texto_com),
                    "cumple_sla": (t_resp_min <= sla_horas * 60) if t_resp_min is not None else None
                })

            likes  = post.get("likes", {}).get("summary", {}).get("total_count", 0)
            shares = post.get("shares", {}).get("count", 0)
            n_com  = post.get("comments", {}).get("summary", {}).get("total_count", 0)
            resultados.append({
                "id": post["id"], "texto": post.get("message", "(Sin texto)"),
                "fecha": post.get("created_time", ""), "likes": likes, "shares": shares,
                "replies_total": n_com, "comentarios": comentarios,
                "url": f"https://www.facebook.com/{post['id'].replace('_', '/posts/')}"
            })
            print(f"{len(comentarios)} comentarios de usuarios")
        return resultados


# ──────────────────────────────────────────────────────────
# MÓDULO INSTAGRAM
# ──────────────────────────────────────────────────────────
class InstagramMonitor:
    BASE = "https://graph.facebook.com/v19.0"

    def __init__(self, user_id: str, access_token: str, ig_username: str):
        self.user_id = user_id
        self.token = access_token
        self.ig_username = ig_username.lstrip("@").lower()

    def _get(self, endpoint: str, params: Dict = None) -> Optional[Dict]:
        params = dict(params or {})
        params["access_token"] = self.token
        try:
            r = requests.get(f"{self.BASE}/{endpoint}", params=params, timeout=15)
            if r.status_code != 200:
                print(f"  ✗ IG API {r.status_code}: {r.text[:300]}")
                return None
            return r.json()
        except Exception as e:
            print(f"  ✗ Conexión IG: {e}")
            return None

    def get_media(self, limit: int) -> List[Dict]:
        d = self._get(f"{self.user_id}/media", {
            "fields": "id,caption,timestamp,like_count,comments_count,permalink",
            "limit": limit
        })
        return d.get("data", []) if d else []

    def get_comments(self, media_id: str) -> List[Dict]:
        d = self._get(f"{media_id}/comments", {
            "fields": "id,username,text,timestamp,replies{id,username,text,timestamp}",
            "limit": 100
        })
        return d.get("data", []) if d else []

    def analizar(self, max_posts: int, sla_horas: float) -> List[Dict]:
        print(f"\n📸 Analizando Instagram (@{self.ig_username}) — SLA: {sla_horas}h...")
        if not self.user_id:
            print("  No configurado. Ver config.json → instagram.user_id")
            return []

        media = self.get_media(max_posts)
        if not media:
            print("  Sin publicaciones o credenciales inválidas.")
            return []

        resultados = []
        for item in media:
            print(f"  → Media {item['id'][-10:]}... ", end="", flush=True)
            raw_comments = self.get_comments(item["id"])

            comentarios = []
            for com in raw_comments:
                if com.get("username", "").lower() == self.ig_username:
                    continue
                replies = com.get("replies", {}).get("data", [])
                respondido = False; t_resp_min = None; txt_resp = None
                for rep in replies:
                    if rep.get("username", "").lower() == self.ig_username:
                        respondido = True
                        tc = parsear_fecha(com["timestamp"])
                        tr = parsear_fecha(rep["timestamp"])
                        if tc and tr:
                            t_resp_min = (tr - tc).total_seconds() / 60
                        txt_resp = rep.get("text", "")
                        break
                texto_com = com.get("text", "")
                comentarios.append({
                    "usuario": f"@{com.get('username', 'usuario')}",
                    "texto": texto_com,
                    "fecha": com.get("timestamp", ""),
                    "respondido": respondido,
                    "tiempo_respuesta_min": t_resp_min,
                    "texto_respuesta": txt_resp,
                    "tema": clasificar_tema(texto_com),
                    "calidad": evaluar_calidad(texto_com, txt_resp or ""),
                    "emocion": detectar_emocion(texto_com),
                    "cumple_sla": (t_resp_min <= sla_horas * 60) if t_resp_min is not None else None
                })

            resultados.append({
                "id": item["id"], "texto": item.get("caption", "(Sin caption)"),
                "fecha": item.get("timestamp", ""), "likes": item.get("like_count", 0),
                "shares": 0, "replies_total": item.get("comments_count", 0),
                "comentarios": comentarios, "url": item.get("permalink", "")
            })
            print(f"{len(comentarios)} comentarios de usuarios")
        return resultados


# ──────────────────────────────────────────────────────────
# MÉTRICAS
# ──────────────────────────────────────────────────────────
def calcular_metricas(publicaciones: List[Dict], sla_h: float) -> Dict:
    total_com  = sum(len(p["comentarios"]) for p in publicaciones)
    respondidos = sum(1 for p in publicaciones for c in p["comentarios"] if c["respondido"])
    tiempos    = [c["tiempo_respuesta_min"] for p in publicaciones for c in p["comentarios"]
                  if c["respondido"] and c["tiempo_respuesta_min"] is not None]
    a_tiempo   = sum(1 for t in tiempos if t <= sla_h * 60)
    calidades  = [c["calidad"]["score"] for p in publicaciones for c in p["comentarios"] if c["respondido"]]

    temas: Dict[str, int] = {}
    emociones: Dict[str, int] = {"Positivo": 0, "Negativo": 0, "Crítico": 0, "Neutro": 0}
    for p in publicaciones:
        for c in p["comentarios"]:
            temas[c["tema"]] = temas.get(c["tema"], 0) + 1
            emociones[c.get("emocion", "Neutro")] = emociones.get(c.get("emocion", "Neutro"), 0) + 1

    return {
        "total_publicaciones": len(publicaciones),
        "total_comentarios": total_com,
        "respondidos": respondidos,
        "sin_respuesta": total_com - respondidos,
        "tasa_respuesta": round(respondidos / total_com * 100, 1) if total_com else 0,
        "tiempo_promedio_min": round(sum(tiempos) / len(tiempos), 1) if tiempos else None,
        "cumplimiento_sla": round(a_tiempo / len(tiempos) * 100, 1) if tiempos else 0,
        "calidad_promedio": round(sum(calidades) / len(calidades), 2) if calidades else 0,
        "resp_sin_resp": sum(1 for s in calidades if s == 0),
        "resp_genericas": sum(1 for s in calidades if s == 1),
        "resp_parciales": sum(1 for s in calidades if s == 2),
        "resp_especificas": sum(1 for s in calidades if s == 3),
        "temas": dict(sorted(temas.items(), key=lambda x: x[1], reverse=True)),
        "emociones": emociones,
    }


# ──────────────────────────────────────────────────────────
# ALERTAS POR CORREO
# ──────────────────────────────────────────────────────────
def enviar_alerta(config: Dict, datos: Dict[str, List[Dict]], reporte_path: str):
    cfg = config.get("alertas", {})
    if not cfg.get("activo", False):
        return

    sin_resp = [
        (plat, pub["url"], com["usuario"], com["texto"], com["fecha"])
        for plat, pubs in datos.items()
        for pub in pubs
        for com in pub["comentarios"]
        if not com["respondido"]
    ]

    umbral = cfg.get("umbral_sin_respuesta", 5)
    if len(sin_resp) < umbral:
        print(f"  ℹ️  Alertas: {len(sin_resp)} sin respuesta (umbral {umbral}). No se envía.")
        return

    print(f"  📧 Enviando alerta: {len(sin_resp)} comentarios sin respuesta...")

    filas = "".join(
        f"<tr><td>{p.title()}</td><td><a href='{u}'>{u[:50]}…</a></td>"
        f"<td>{html_lib.escape(usr)}</td><td>{html_lib.escape(txt[:100])}…</td><td>{f}</td></tr>"
        for p, u, usr, txt, f in sin_resp[:50]
    )

    cuerpo = f"""
    <html><body style="font-family:Arial,sans-serif">
    <h2 style="color:#c0392b">⚠️ Alerta ICETEX Social Monitor</h2>
    <p>Se detectaron <strong>{len(sin_resp)} comentarios sin respuesta</strong>
       (umbral configurado: {umbral}).</p>
    <table border="1" cellpadding="6" cellspacing="0" style="border-collapse:collapse;width:100%">
      <thead style="background:#f0f4ff">
        <tr><th>Plataforma</th><th>Publicación</th><th>Usuario</th>
            <th>Comentario</th><th>Fecha</th></tr>
      </thead>
      <tbody>{filas}</tbody>
    </table>
    <p style="color:#777;font-size:0.85rem;margin-top:20px">
      Reporte completo adjunto. Generado el {datetime.now().strftime('%d/%m/%Y %H:%M')}.
    </p>
    </body></html>
    """

    msg = MIMEMultipart()
    msg["From"]    = cfg["usuario"]
    msg["To"]      = ", ".join(cfg["destinatarios"])
    msg["Subject"] = f"⚠️ ICETEX Monitor — {len(sin_resp)} comentarios sin respuesta"
    msg.attach(MIMEText(cuerpo, "html", "utf-8"))

    if os.path.exists(reporte_path):
        with open(reporte_path, "rb") as f:
            part = MIMEBase("application", "octet-stream")
            part.set_payload(f.read())
        encoders.encode_base64(part)
        part.add_header("Content-Disposition", f'attachment; filename="{os.path.basename(reporte_path)}"')
        msg.attach(part)

    try:
        with smtplib.SMTP(cfg["smtp_host"], cfg["smtp_port"]) as server:
            server.starttls()
            server.login(cfg["usuario"], cfg["password"])
            server.sendmail(cfg["usuario"], cfg["destinatarios"], msg.as_string())
        print("  ✅ Alerta enviada.")
    except Exception as e:
        print(f"  ✗ Error enviando alerta: {e}")


# ──────────────────────────────────────────────────────────
# EXPORTACIÓN EXCEL
# ──────────────────────────────────────────────────────────
def exportar_excel(datos: Dict[str, List[Dict]], metricas_plat: Dict, mg: Dict,
                   sla_config: Dict, output_path: str):
    if not EXCEL_OK:
        print("  ⚠️  openpyxl no disponible. Instala con: pip install openpyxl")
        return

    wb = openpyxl.Workbook()

    # Colores corporativos
    AZUL_ICETEX  = "0A4B8C"
    AZUL_CLARO   = "1976D2"
    VERDE        = "27AE60"
    ROJO         = "E74C3C"
    NARANJA      = "F39C12"
    GRIS_HEADER  = "F0F4FF"
    BLANCO       = "FFFFFF"

    def hdr_fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def bold_font(color="000000", size=11):
        return Font(bold=True, color=color, size=size)

    def center():
        return Alignment(horizontal="center", vertical="center", wrap_text=True)

    def left():
        return Alignment(horizontal="left", vertical="center", wrap_text=True)

    thin = Side(border_style="thin", color="CCCCCC")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Hoja 1: Resumen
    ws = wb.active
    ws.title = "Resumen"
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:G1")
    ws["A1"] = "ICETEX — Monitor Social · Resumen Ejecutivo"
    ws["A1"].font = Font(bold=True, color=BLANCO, size=14)
    ws["A1"].fill = hdr_fill(AZUL_ICETEX)
    ws["A1"].alignment = center()
    ws.row_dimensions[1].height = 36

    ws["A2"] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws["A2"].font = Font(italic=True, color="888888", size=9)
    ws.merge_cells("A2:G2")
    ws.row_dimensions[2].height = 16

    # Encabezado tabla resumen
    headers_res = ["Plataforma", "Publicaciones", "Comentarios", "Respondidos",
                   "Tasa Resp. %", "Tiempo Prom.", "Cumpl. SLA %"]
    for col, h in enumerate(headers_res, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = bold_font(BLANCO)
        cell.fill = hdr_fill(AZUL_CLARO)
        cell.alignment = center()
        cell.border = border_thin
    ws.row_dimensions[4].height = 22

    PLAT_NOMBRES = {"twitter": "X (Twitter)", "facebook": "Facebook", "instagram": "Instagram"}
    row = 5
    for plat, m in metricas_plat.items():
        sla_h = sla_config.get(plat, 24)
        vals = [
            PLAT_NOMBRES.get(plat, plat),
            m["total_publicaciones"], m["total_comentarios"], m["respondidos"],
            m["tasa_respuesta"], fmt_tiempo(m["tiempo_promedio_min"]), m["cumplimiento_sla"]
        ]
        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=v)
            cell.alignment = center()
            cell.border = border_thin
            if col == 5:  # tasa respuesta
                pct = float(v) if isinstance(v, (int, float)) else 0
                cell.fill = hdr_fill(VERDE if pct >= 85 else (NARANJA if pct >= 60 else ROJO))
                cell.font = bold_font(BLANCO)
            elif col == 7:  # SLA
                pct = float(v) if isinstance(v, (int, float)) else 0
                cell.fill = hdr_fill(VERDE if pct >= 85 else (NARANJA if pct >= 60 else ROJO))
                cell.font = bold_font(BLANCO)
        row += 1

    # Fila GLOBAL
    if mg:
        glb_vals = [
            "GLOBAL",
            mg["total_publicaciones"], mg["total_comentarios"], mg["respondidos"],
            mg["tasa_respuesta"], fmt_tiempo(mg["tiempo_promedio_min"]), mg["cumplimiento_sla"]
        ]
        for col, v in enumerate(glb_vals, 1):
            cell = ws.cell(row=row, column=col, value=v)
            cell.font = bold_font()
            cell.fill = hdr_fill(GRIS_HEADER)
            cell.alignment = center()
            cell.border = border_thin
        row += 1

    # Columnas resumen
    for w, col in zip([22, 14, 13, 13, 12, 13, 12], range(1, 8)):
        ws.column_dimensions[get_column_letter(col)].width = w

    # ── Hojas de detalle por plataforma
    for plat, pubs in datos.items():
        if not pubs:
            continue
        ws2 = wb.create_sheet(title=PLAT_NOMBRES.get(plat, plat))
        ws2.sheet_view.showGridLines = False

        COLOR_PLAT = {"twitter": "14171A", "facebook": "1877F2", "instagram": "C13584"}
        color_plat = COLOR_PLAT.get(plat, AZUL_ICETEX)

        sla_h = sla_config.get(plat, 24)
        ws2.merge_cells("A1:J1")
        ws2["A1"] = f"{PLAT_NOMBRES.get(plat, plat)} — SLA: {sla_h}h"
        ws2["A1"].font = Font(bold=True, color=BLANCO, size=12)
        ws2["A1"].fill = hdr_fill(color_plat)
        ws2["A1"].alignment = center()
        ws2.row_dimensions[1].height = 28

        hdrs = ["Fecha", "Publicación", "Usuario", "Comentario", "Tema",
                "Respondido", "Tiempo Resp.", "Cumple SLA", "Calidad", "Emoción", "Texto Respuesta"]
        for col, h in enumerate(hdrs, 1):
            cell = ws2.cell(row=2, column=col, value=h)
            cell.font = bold_font(BLANCO)
            cell.fill = hdr_fill(AZUL_CLARO)
            cell.alignment = center()
            cell.border = border_thin
        ws2.row_dimensions[2].height = 20

        fila = 3
        for pub in pubs:
            for com in pub["comentarios"]:
                respondido = com["respondido"]
                t_resp     = com["tiempo_respuesta_min"]
                cumple     = com.get("cumple_sla")
                calidad    = com["calidad"]

                vals = [
                    com["fecha"][:10] if com["fecha"] else "",
                    pub["texto"][:80],
                    com["usuario"],
                    com["texto"][:120],
                    com["tema"],
                    "Sí" if respondido else "No",
                    fmt_tiempo(t_resp) if respondido else "—",
                    ("Sí" if cumple else "No") if cumple is not None else "—",
                    calidad["nivel"],
                    com.get("emocion", "Neutro"),
                    (com.get("texto_respuesta") or "")[:150]
                ]
                for col, v in enumerate(vals, 1):
                    cell = ws2.cell(row=fila, column=col, value=v)
                    cell.alignment = left()
                    cell.border = border_thin

                    if col == 6:  # Respondido
                        cell.font = Font(bold=True,
                                         color=VERDE if respondido else ROJO)
                        cell.alignment = center()
                    if col == 8:  # Cumple SLA
                        cell.alignment = center()
                        if cumple is True:
                            cell.font = Font(bold=True, color=VERDE)
                        elif cumple is False:
                            cell.font = Font(bold=True, color=ROJO)
                    if col == 9:  # Calidad
                        c_map = {0: ROJO, 1: NARANJA, 2: AZUL_CLARO, 3: VERDE}
                        cell.font = Font(bold=True, color=c_map.get(calidad["score"], "000000"))
                        cell.alignment = center()
                    if col == 10:  # Emoción
                        e_map = {"Positivo": VERDE, "Negativo": NARANJA, "Crítico": ROJO, "Neutro": "7F8C8D"}
                        cell.font = Font(bold=True, color=e_map.get(com.get("emocion", "Neutro"), "000000"))
                        cell.alignment = center()
                fila += 1

        ws2.freeze_panes = "A3"
        for w, col in zip([12, 30, 18, 40, 22, 10, 12, 11, 12, 12, 45], range(1, 12)):
            ws2.column_dimensions[get_column_letter(col)].width = w

    try:
        wb.save(output_path)
        print(f"📊 Excel guardado en:\n   {output_path}")
    except Exception as e:
        print(f"  ✗ Error guardando Excel: {e}")


# ──────────────────────────────────────────────────────────
# REPORTE HTML
# ──────────────────────────────────────────────────────────
def c_tasa(pct: float) -> str:
    return "#27ae60" if pct >= 85 else ("#f39c12" if pct >= 60 else "#e74c3c")

def c_tiempo(minutos: Optional[float], sla_min: float) -> str:
    if minutos is None: return "#95a5a6"
    if minutos <= sla_min * 0.5: return "#27ae60"
    if minutos <= sla_min: return "#f39c12"
    return "#e74c3c"

def badge_calidad(score: int) -> str:
    mapa = {
        0: ("Sin respuesta", "#e74c3c"),
        1: ("Genérica",      "#e67e22"),
        2: ("Parcial",       "#2980b9"),
        3: ("Específica",    "#27ae60")
    }
    label, color = mapa.get(score, ("?", "#95a5a6"))
    return (f'<span style="background:{color};color:#fff;padding:2px 9px;'
            f'border-radius:10px;font-size:0.73rem;font-weight:700;white-space:nowrap">'
            f'{label}</span>')


def badge_emocion(emocion: str) -> str:
    mapa = {
        "Positivo": ("#27ae60", "😊"),
        "Negativo": ("#e67e22", "😞"),
        "Crítico":  ("#e74c3c", "😡"),
        "Neutro":   ("#7f8c8d", "😐"),
    }
    color, emoji = mapa.get(emocion, ("#95a5a6", "❓"))
    return (f'<span style="background:{color};color:#fff;padding:2px 9px;'
            f'border-radius:10px;font-size:0.73rem;font-weight:700;white-space:nowrap">'
            f'{emoji} {emocion}</span>')


def generar_reporte(datos: Dict[str, List[Dict]], config: Dict,
                    sla_config: Dict, metricas_plat: Dict, mg: Dict) -> str:
    fecha      = datetime.now().strftime("%d/%m/%Y %H:%M")
    fecha_safe = datetime.now().strftime("%Y%m%d_%H%M")
    datos_json = json.dumps(datos, ensure_ascii=False, default=str)

    PLAT = {
        "twitter":   {"nombre": "X (Twitter)", "color": "#14171A", "icon": "𝕏"},
        "facebook":  {"nombre": "Facebook",    "color": "#1877F2", "icon": "f"},
        "instagram": {"nombre": "Instagram",   "color": "#C13584", "icon": "◉"},
    }

    # ── KPI global
    if mg:
        tr = mg.get("tasa_respuesta", 0)
        tp = mg.get("tiempo_promedio_min")
        sla_pct = mg.get("cumplimiento_sla", 0)
        cal = mg.get("calidad_promedio", 0)
        sla_global = max(sla_config.values()) if sla_config else 24
        cards_html = f"""
        <div class="kpi-grid">
          <div class="kpi"><div class="kpi-l">Total Comentarios</div>
            <div class="kpi-v">{mg.get('total_comentarios',0)}</div>
            <div class="kpi-s">{mg.get('total_publicaciones',0)} publicaciones analizadas</div></div>
          <div class="kpi"><div class="kpi-l">Tasa de Respuesta</div>
            <div class="kpi-v" style="color:{c_tasa(tr)}">{tr}%</div>
            <div class="kpi-s">{mg.get('respondidos',0)} respondidos · {mg.get('sin_respuesta',0)} sin respuesta</div></div>
          <div class="kpi"><div class="kpi-l">Tiempo Promedio</div>
            <div class="kpi-v" style="color:{c_tiempo(tp, sla_global*60)}">{fmt_tiempo(tp)}</div>
            <div class="kpi-s">SLA máx. configurado: {sla_global}h</div></div>
          <div class="kpi"><div class="kpi-l">Cumplimiento SLA</div>
            <div class="kpi-v" style="color:{c_tasa(sla_pct)}">{sla_pct}%</div>
            <div class="kpi-s">Respuestas dentro del SLA</div></div>
          <div class="kpi"><div class="kpi-l">Calidad Respuestas</div>
            <div class="kpi-v" style="color:{c_tasa(cal/3*100)}">{cal:.1f}/3</div>
            <div class="kpi-s">{mg.get('resp_genericas',0)} genéricas · {mg.get('resp_especificas',0)} específicas</div></div>
        </div>"""
    else:
        cards_html = '<p style="color:#888;margin-bottom:32px">Sin datos. Verifica credenciales en config.json</p>'

    # ── Datos para gráficas
    plat_nombres  = [PLAT[k]["nombre"] for k in datos if datos[k]]
    plat_tasa     = [metricas_plat.get(k, {}).get("tasa_respuesta", 0) for k in datos if datos[k]]
    plat_sla      = [metricas_plat.get(k, {}).get("cumplimiento_sla", 0) for k in datos if datos[k]]
    plat_colors   = [PLAT[k]["color"] for k in datos if datos[k]]

    temas_global: Dict[str, int] = {}
    for m in metricas_plat.values():
        for t, n in m.get("temas", {}).items():
            temas_global[t] = temas_global.get(t, 0) + n
    temas_sorted = sorted(temas_global.items(), key=lambda x: x[1], reverse=True)
    temas_labels = [t for t, _ in temas_sorted]
    temas_values = [n for _, n in temas_sorted]

    cal_labels = ["Sin respuesta", "Genérica", "Parcial", "Específica"]
    cal_counts = [mg.get("resp_sin_resp", 0), mg.get("resp_genericas", 0),
                  mg.get("resp_parciales", 0), mg.get("resp_especificas", 0)] if mg else [0, 0, 0, 0]
    cal_colors_js = ["#e74c3c", "#e67e22", "#2980b9", "#27ae60"]

    emociones_global: Dict[str, int] = {"Positivo": 0, "Negativo": 0, "Crítico": 0, "Neutro": 0}
    for m in metricas_plat.values():
        for e, n in m.get("emociones", {}).items():
            emociones_global[e] = emociones_global.get(e, 0) + n
    em_labels = ["Positivo", "Negativo", "Crítico", "Neutro"]
    em_counts  = [emociones_global.get(e, 0) for e in em_labels]
    em_colors  = ["#27ae60", "#e67e22", "#e74c3c", "#7f8c8d"]

    charts_section = f"""
    <div class="charts-grid">
      <div class="chart-card">
        <div class="chart-title">Tasa de Respuesta por Plataforma</div>
        <canvas id="chartTasa"></canvas>
      </div>
      <div class="chart-card">
        <div class="chart-title">Calidad de Respuestas (Global)</div>
        <canvas id="chartCalidad"></canvas>
      </div>
      <div class="chart-card chart-wide">
        <div class="chart-title">Temas más frecuentes (Global)</div>
        <canvas id="chartTemas"></canvas>
      </div>
      <div class="chart-card">
        <div class="chart-title">Cumplimiento SLA por Plataforma</div>
        <canvas id="chartSLA"></canvas>
      </div>
      <div class="chart-card">
        <div class="chart-title">Distribución de Emociones (Global)</div>
        <canvas id="chartEmociones"></canvas>
      </div>
    </div>
    <script>
    const chartDefaults = {{ responsive: true, maintainAspectRatio: true,
      plugins: {{ legend: {{ labels: {{ font: {{ size: 12, family: "'Segoe UI', Arial" }} }} }} }} }};

    new Chart(document.getElementById('chartTasa'), {{
      type: 'bar',
      data: {{
        labels: {json.dumps(plat_nombres)},
        datasets: [{{ label: 'Tasa de Respuesta (%)',
          data: {json.dumps(plat_tasa)},
          backgroundColor: {json.dumps(plat_colors)},
          borderRadius: 6 }}]
      }},
      options: {{ ...chartDefaults, scales: {{
        y: {{ min:0, max:100, ticks: {{ callback: v => v+'%' }} }},
        x: {{ grid: {{ display: false }} }}
      }}, plugins: {{ ...chartDefaults.plugins, legend: {{ display: false }} }} }}
    }});

    new Chart(document.getElementById('chartCalidad'), {{
      type: 'doughnut',
      data: {{
        labels: {json.dumps(cal_labels)},
        datasets: [{{ data: {json.dumps(cal_counts)},
          backgroundColor: {json.dumps(cal_colors_js)},
          borderWidth: 2, borderColor: '#fff' }}]
      }},
      options: {{ ...chartDefaults, cutout: '62%',
        plugins: {{ ...chartDefaults.plugins,
          legend: {{ position: 'bottom', labels: {{ padding: 14, font: {{ size: 11 }} }} }} }} }}
    }});

    new Chart(document.getElementById('chartTemas'), {{
      type: 'bar',
      data: {{
        labels: {json.dumps(temas_labels)},
        datasets: [{{ label: 'Comentarios',
          data: {json.dumps(temas_values)},
          backgroundColor: '#1976d2cc', borderRadius: 4 }}]
      }},
      options: {{ ...chartDefaults, indexAxis: 'y',
        scales: {{ x: {{ beginAtZero: true }}, y: {{ grid: {{ display: false }} }} }},
        plugins: {{ ...chartDefaults.plugins, legend: {{ display: false }} }} }}
    }});

    new Chart(document.getElementById('chartSLA'), {{
      type: 'bar',
      data: {{
        labels: {json.dumps(plat_nombres)},
        datasets: [{{ label: 'Cumplimiento SLA (%)',
          data: {json.dumps(plat_sla)},
          backgroundColor: {json.dumps(plat_colors)},
          borderRadius: 6 }}]
      }},
      options: {{ ...chartDefaults, scales: {{
        y: {{ min:0, max:100, ticks: {{ callback: v => v+'%' }} }},
        x: {{ grid: {{ display: false }} }}
      }}, plugins: {{ ...chartDefaults.plugins, legend: {{ display: false }} }} }}
    }});

    new Chart(document.getElementById('chartEmociones'), {{
      type: 'doughnut',
      data: {{
        labels: {json.dumps(em_labels)},
        datasets: [{{ data: {json.dumps(em_counts)},
          backgroundColor: {json.dumps(em_colors)},
          borderWidth: 2, borderColor: '#fff' }}]
      }},
      options: {{ ...chartDefaults, cutout: '62%',
        plugins: {{ ...chartDefaults.plugins,
          legend: {{ position: 'bottom', labels: {{ padding: 14, font: {{ size: 11 }} }} }} }} }}
    }});
    </script>"""

    # ── Secciones por plataforma
    plat_html = ""
    for plat, pubs in datos.items():
        if not pubs:
            continue
        info  = PLAT.get(plat, {"nombre": plat, "color": "#555", "icon": "?"})
        m     = metricas_plat.get(plat, {})
        sla_h = sla_config.get(plat, 24)
        tr_val   = m.get("tasa_respuesta", 0)
        tp_val   = m.get("tiempo_promedio_min")
        sla_val  = m.get("cumplimiento_sla", 0)

        rows = ""
        for pub in pubs:
            pub_texto = html_lib.escape(pub["texto"][:70]) + ("…" if len(pub["texto"]) > 70 else "")
            pub_url   = pub.get("url", "#")
            for com in pub["comentarios"]:
                txt_r   = com.get("texto_respuesta") or ""
                tr_min  = com.get("tiempo_respuesta_min")
                cumple  = com.get("cumple_sla")
                if cumple is True:
                    sla_badge = '<span style="color:#27ae60;font-weight:700">✓ Sí</span>'
                elif cumple is False:
                    sla_badge = '<span style="color:#e74c3c;font-weight:700">✗ No</span>'
                else:
                    sla_badge = '<span style="color:#aaa">—</span>'

                rows += f"""<tr>
                  <td><a href="{pub_url}" target="_blank" class="pub-link"
                      title="{html_lib.escape(pub['texto'])}">{pub_texto}</a></td>
                  <td class="nowrap">{html_lib.escape(str(com['usuario']))}</td>
                  <td title="{html_lib.escape(com['texto'])}">{html_lib.escape(com['texto'][:90])}{'…' if len(com['texto'])>90 else ''}</td>
                  <td><span class="tag">{com['tema']}</span></td>
                  <td class="nowrap" style="color:{'#27ae60' if com['respondido'] else '#e74c3c'};font-weight:700">
                    {'✓ Sí' if com['respondido'] else '✗ No'}</td>
                  <td class="nowrap" style="color:{c_tiempo(tr_min, sla_h*60) if com['respondido'] else '#aaa'}">
                    {fmt_tiempo(tr_min) if com['respondido'] else '—'}</td>
                  <td class="nowrap">{sla_badge}</td>
                  <td>{badge_calidad(com['calidad']['score'])}</td>
                  <td>{badge_emocion(com.get('emocion','Neutro'))}</td>
                  <td title="{html_lib.escape(txt_r)}">{html_lib.escape(txt_r[:80])}{'…' if len(txt_r)>80 else ''}</td>
                </tr>"""

        temas_chips = " ".join(
            f'<span class="chip">{t} <b>{n}</b></span>'
            for t, n in m.get("temas", {}).items()
        )

        plat_html += f"""
        <section class="plat-section">
          <div class="plat-hdr" style="background:{info['color']}">
            <span class="plat-icon">{info['icon']}</span>
            <span class="plat-name">{info['nombre']}</span>
            <div class="plat-kpis">
              <span class="pkpi">📝 {m.get('total_publicaciones',0)} pubs.</span>
              <span class="pkpi">💬 {m.get('total_comentarios',0)} comentarios</span>
              <span class="pkpi" style="color:{'#a8ffb0' if tr_val>=80 else '#ffd080' if tr_val>=60 else '#ff9090'}">
                ↩ {tr_val}% resp.</span>
              <span class="pkpi">⏱ {fmt_tiempo(tp_val)} prom.</span>
              <span class="pkpi">SLA ({sla_h}h): {sla_val}%</span>
            </div>
          </div>
          <div class="temas-bar">{temas_chips}</div>
          <div class="tbl-wrap">
            <table>
              <thead><tr>
                <th>Publicación</th><th>Usuario</th><th>Comentario</th>
                <th>Tema</th><th>Respondido</th><th>Tiempo Resp.</th>
                <th>Cumple SLA</th><th>Calidad</th><th>Emoción</th><th>Texto Respuesta</th>
              </tr></thead>
              <tbody>{rows or '<tr><td colspan="10" class="empty">Sin comentarios en este período</td></tr>'}</tbody>
            </table>
          </div>
        </section>"""

    return f"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Monitor Social ICETEX — {fecha}</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"></script>
<script src="https://cdn.jsdelivr.net/npm/xlsx@0.18.5/dist/xlsx.full.min.js"></script>
<style>
*{{box-sizing:border-box;margin:0;padding:0}}
body{{font-family:'Segoe UI',Arial,sans-serif;background:#f0f2f5;color:#1a1a2e}}
a{{color:inherit;text-decoration:none}}

.app-hdr{{background:linear-gradient(135deg,#0a4b8c,#1565c0 60%,#1976d2);
  color:#fff;padding:24px 40px;display:flex;align-items:center;gap:20px;
  box-shadow:0 4px 20px rgba(0,0,0,.2)}}
.app-hdr h1{{font-size:1.55rem;font-weight:800;letter-spacing:-.5px}}
.app-hdr p{{font-size:.88rem;opacity:.85;margin-top:3px}}
.hdr-badge{{background:rgba(255,255,255,.18);border-radius:10px;padding:10px 18px;
  margin-left:auto;font-size:.83rem;text-align:center;line-height:1.9}}
.btn-excel{{display:inline-block;margin-top:10px;background:#217346;color:#fff;
  border:none;border-radius:8px;padding:8px 18px;font-size:.83rem;font-weight:700;
  cursor:pointer;text-decoration:none;transition:background .2s}}
.btn-excel:hover{{background:#185c37}}

.main{{max-width:1500px;margin:0 auto;padding:28px 24px}}
.sec-title{{font-size:1.05rem;font-weight:800;color:#0a4b8c;margin:28px 0 16px;
  padding-bottom:8px;border-bottom:3px solid #1976d2}}

.kpi-grid{{display:grid;grid-template-columns:repeat(auto-fit,minmax(190px,1fr));gap:16px;margin-bottom:8px}}
.kpi{{background:#fff;border-radius:14px;padding:20px 18px;box-shadow:0 2px 12px rgba(0,0,0,.06);
  border-top:4px solid #1976d2}}
.kpi-l{{font-size:.75rem;font-weight:700;color:#666;text-transform:uppercase;letter-spacing:.5px}}
.kpi-v{{font-size:2rem;font-weight:900;margin:8px 0 4px;line-height:1}}
.kpi-s{{font-size:.75rem;color:#888}}

.charts-grid{{display:grid;grid-template-columns:repeat(2,1fr);gap:20px;margin-bottom:8px}}
.chart-wide{{grid-column:1/-1}}
.chart-card{{background:#fff;border-radius:14px;padding:20px 22px;
  box-shadow:0 2px 12px rgba(0,0,0,.06)}}
.chart-title{{font-size:.85rem;font-weight:700;color:#0a4b8c;margin-bottom:14px}}

.plat-section{{background:#fff;border-radius:14px;margin-bottom:28px;
  box-shadow:0 2px 12px rgba(0,0,0,.07);overflow:hidden}}
.plat-hdr{{display:flex;align-items:center;gap:14px;padding:15px 22px;color:#fff;flex-wrap:wrap}}
.plat-icon{{font-size:1.4rem;font-weight:900;width:34px;text-align:center}}
.plat-name{{font-size:1.05rem;font-weight:700}}
.plat-kpis{{display:flex;gap:10px;margin-left:auto;flex-wrap:wrap}}
.pkpi{{background:rgba(255,255,255,.18);border-radius:20px;padding:4px 12px;
  font-size:.8rem;font-weight:600}}

.temas-bar{{padding:10px 22px;background:#f8f9fa;border-bottom:1px solid #eee;
  display:flex;flex-wrap:wrap;gap:8px;align-items:center;font-size:.78rem}}
.temas-bar::before{{content:"Temas:";font-weight:700;color:#666;margin-right:4px}}
.chip{{background:#e8f0fe;color:#1565c0;border-radius:20px;padding:3px 11px;font-size:.77rem}}
.chip b{{color:#0a4b8c}}

.tbl-wrap{{overflow-x:auto;padding:14px 22px 20px}}
table{{width:100%;border-collapse:collapse;font-size:.81rem}}
thead th{{background:#f0f4ff;color:#0a4b8c;font-weight:700;padding:10px 11px;
  text-align:left;border-bottom:2px solid #d0deff;white-space:nowrap}}
tbody tr{{border-bottom:1px solid #f0f0f0;transition:background .12s}}
tbody tr:hover{{background:#f7f9ff}}
td{{padding:8px 11px;vertical-align:middle;max-width:240px;
  overflow:hidden;text-overflow:ellipsis;white-space:nowrap}}
.nowrap{{white-space:nowrap}}
.tag{{background:#e8f0fe;color:#1565c0;border-radius:12px;padding:2px 8px;
  font-size:.72rem;font-weight:700;white-space:nowrap}}
.pub-link:hover{{text-decoration:underline;color:#1976d2}}
.empty{{text-align:center;color:#aaa;padding:20px!important}}

.footer{{text-align:center;padding:22px;color:#999;font-size:.78rem;margin-top:8px}}
.footer a{{color:#1976d2}}

@media(max-width:900px){{
  .charts-grid{{grid-template-columns:1fr}}
  .chart-wide{{grid-column:auto}}
  .app-hdr{{padding:16px;flex-wrap:wrap}}
  .hdr-badge{{margin-left:0;width:100%}}
  .kpi-grid{{grid-template-columns:1fr 1fr}}
  .plat-kpis{{margin-left:0;width:100%}}
}}
</style>
</head>
<body>

<script>
const _datos = {datos_json};
function exportarExcel() {{
  if (typeof XLSX === 'undefined') {{ alert('Requiere conexión a internet para exportar.'); return; }}
  const wb = XLSX.utils.book_new();
  const plats = {{'facebook':'Facebook','instagram':'Instagram','twitter':'Twitter'}};
  let resumen = [['Plataforma','Publicaciones','Comentarios','Respondidos','Sin Respuesta','Tasa Resp %','T.Promedio (min)']];
  Object.entries(_datos).forEach(([plat, pubs]) => {{
    if (!pubs.length) return;
    const nombre = plats[plat] || plat;
    let coms=0,resp=0,tiempos=[];
    pubs.forEach(p=>{{ p.comentarios.forEach(c=>{{ coms++; if(c.respondido){{ resp++; if(c.tiempo_respuesta_min!=null) tiempos.push(c.tiempo_respuesta_min); }} }}); }});
    const tasa = coms ? (resp/coms*100).toFixed(1) : 0;
    const tprom = tiempos.length ? (tiempos.reduce((a,b)=>a+b,0)/tiempos.length).toFixed(0) : '-';
    resumen.push([nombre, pubs.length, coms, resp, coms-resp, tasa, tprom]);
    let rows = [['Fecha Post','Publicación','Usuario','Comentario','Tema','Respondido','T.Respuesta (min)','Cumple SLA','Calidad','Texto Respuesta']];
    pubs.forEach(p=>{{ p.comentarios.forEach(c=>{{ rows.push([
      p.fecha?p.fecha.substring(0,10):'', p.texto?p.texto.substring(0,80):'',
      c.usuario, c.texto, c.tema, c.respondido?'Sí':'No',
      c.tiempo_respuesta_min!=null?c.tiempo_respuesta_min:'-',
      c.cumple_sla===true?'Sí':c.cumple_sla===false?'No':'-',
      c.calidad?c.calidad.nivel:'', c.texto_respuesta||''
    ]); }}); }});
    XLSX.utils.book_append_sheet(wb, XLSX.utils.aoa_to_sheet(rows), nombre);
  }});
  XLSX.utils.book_append_sheet(wb, XLSX.utils.aoa_to_sheet(resumen), 'Resumen');
  XLSX.writeFile(wb, 'reporte_social_{fecha_safe}.xlsx');
}}
</script>

<header class="app-hdr">
  <div>
    <h1>Monitor Social ICETEX</h1>
    <p>Análisis de interacciones · Facebook · X (Twitter) · Instagram</p>
  </div>
  <div class="hdr-badge">
    📅 {fecha}<br>
    🐦 SLA X: {sla_config.get('twitter',4)}h &nbsp;·&nbsp;
    📘 FB: {sla_config.get('facebook',8)}h &nbsp;·&nbsp;
    📸 IG: {sla_config.get('instagram',12)}h<br>
    <button class="btn-excel" onclick="exportarExcel()">📥 Descargar Excel</button>
  </div>
</header>

<main class="main">
  <div class="sec-title">📊 Resumen Global</div>
  {cards_html}

  <div class="sec-title">📈 Gráficas</div>
  {charts_section}

  <div class="sec-title">🔍 Detalle por Plataforma</div>
  {plat_html or '<p style="color:#888;padding:16px 0">Sin datos por plataforma. Verifica config.json</p>'}
</main>

<footer class="footer">
  Generado el {fecha} · ICETEX Social Monitor v2.0 ·
  <span style="color:#aaa">Las gráficas requieren conexión a internet (Chart.js CDN)</span>
</footer>
</body>
</html>"""


# ──────────────────────────────────────────────────────────
# DATOS DEMO (sin APIs)
# ──────────────────────────────────────────────────────────
def generar_datos_demo(page_name: str = "MiPagina", ig_user: str = "mi_instagram") -> Dict[str, List[Dict]]:
    """Genera publicaciones y comentarios simulados para probar el reporte completo."""
    ahora = datetime.now(timezone.utc)

    def ts(horas_atras: float) -> str:
        return (ahora - timedelta(hours=horas_atras)).isoformat()

    def com(usuario, texto, respondido, t_resp_min, txt_resp, tema=None):
        t = tema or clasificar_tema(texto)
        return {
            "usuario": usuario,
            "texto": texto,
            "fecha": ts(t_resp_min / 60 + 2 if t_resp_min else 5),
            "respondido": respondido,
            "tiempo_respuesta_min": t_resp_min,
            "texto_respuesta": txt_resp,
            "tema": t,
            "calidad": evaluar_calidad(texto, txt_resp or ""),
            "emocion": detectar_emocion(texto),
            "cumple_sla": None if t_resp_min is None else t_resp_min <= 8 * 60,
        }

    fb_posts = [
        {
            "id": "123456_001", "texto": "Conoce nuestras opciones de crédito educativo para el segundo semestre 2026. ¡Aplica ahora!",
            "fecha": ts(48), "likes": 312, "shares": 45, "replies_total": 6, "url": f"https://www.facebook.com/{page_name}",
            "comentarios": [
                com("@juan.perez", "¿Cuánto tiempo demora el desembolso del crédito ACCES?", True, 95,
                    "Hola Juan, el desembolso ACCES tarda entre 8 y 15 días hábiles una vez aprobado tu estudio socioeconómico. Puedes hacer seguimiento en icetex.gov.co"),
                com("@maria.garcia", "Ya llevo 3 semanas esperando respuesta de mi solicitud y nada", False, None, None),
                com("@carlos.ruiz", "¿El crédito cubre también matrícula de posgrado?", True, 210,
                    "Hola Carlos, sí aplica para posgrados en instituciones acreditadas. Revisa los requisitos en icetex.gov.co/creditos"),
                com("@sofia.m", "Terrible el servicio, me cancelaron sin explicación", False, None, None),
            ],
        },
        {
            "id": "123456_002", "texto": "Convocatoria becas Generación E 2026-2 abierta. Consulta requisitos en el enlace.",
            "fecha": ts(24), "likes": 528, "shares": 130, "replies_total": 8, "url": f"https://www.facebook.com/{page_name}",
            "comentarios": [
                com("@andrea.lopez", "¿Cuál es el puntaje mínimo de Sisbén para la beca?", True, 55,
                    "Hola Andrea, para Generación E necesitas Sisbén IV entre A1 y D20. Revisa tu grupo en sisben.gov.co"),
                com("@luis.torres", "¿Las becas Generación E aplican para universidades privadas?", True, 480,
                    "Luis, las becas aplican solo para IES públicas acreditadas de alta calidad. Consulta el listado en nuestra web."),
                com("@paula.v", "Ya perdí la beca porque ICETEX no me notificó a tiempo", False, None, None),
                com("@jorge.h", "¿Cómo renuevo la beca para el siguiente semestre?", True, 38,
                    "Hola Jorge, la renovación se hace desde el portal icetex.gov.co entre el 1 y el 15 del mes de inscripción. Solo necesitas paz y salvo académico."),
            ],
        },
        {
            "id": "123456_003", "texto": "Recuerda que puedes hacer tus pagos de cuotas en línea, sin filas. Ingresa a Mi Portal ICETEX.",
            "fecha": ts(6), "likes": 189, "shares": 22, "replies_total": 4, "url": f"https://www.facebook.com/{page_name}",
            "comentarios": [
                com("@diana.c", "El portal lleva 2 días caído, no puedo pagar mi cuota", True, 25,
                    "Hola Diana, ya estamos trabajando en el restablecimiento. El pago puedes hacerlo por PSE o en puntos Efecty sin recargo por inconveniente técnico."),
                com("@roberto.b", "¿Hay algún beneficio si pago antes del vencimiento?", True, 312,
                    "Roberto, actualmente no contamos con descuento por pago anticipado, pero evitas intereses de mora. Te recomendamos domiciliar el pago."),
            ],
        },
    ]

    ig_media = [
        {
            "id": "IG_001", "texto": "Nueva convocatoria becas internacionales 2026 disponible. ¡No dejes pasar esta oportunidad! #BecasICETEX #EducacionSuperior",
            "fecha": ts(36), "likes": 743, "shares": 0, "replies_total": 5, "url": f"https://www.instagram.com/{ig_user}/",
            "comentarios": [
                com(f"@{ig_user}", "texto", False, None, None),  # propio, se filtra
                com("@valentina_r", "¿Cuáles universidades del exterior están en convenio?", True, 68,
                    "Hola @valentina_r, tenemos convenios con más de 40 universidades en Europa y Latinoamérica. Consulta el directorio completo en icetex.gov.co/internacional"),
                com("@felipe_mm", "Apliqué hace un mes y no tengo respuesta todavía", False, None, None),
                com("@natalia_pg", "¿Esto incluye maestrías o solo pregrado?", True, 145,
                    "Hola @natalia_pg, aplica tanto para pregrado como para maestría y doctorado en instituciones reconocidas."),
            ],
        },
        {
            "id": "IG_002", "texto": "Beneficiarios ICETEX: recuerda activar tu condonación de intereses si cumples los requisitos de rendimiento académico.",
            "fecha": ts(12), "likes": 415, "shares": 0, "replies_total": 3, "url": f"https://www.instagram.com/{ig_user}/",
            "comentarios": [
                com("@camilo_rb", "¿Qué promedio necesito para la condonación?", True, 30,
                    "Hola @camilo_rb, necesitas promedio mínimo de 3.8 sobre 5.0 y haber cursado al menos el 80% de los créditos del semestre."),
                com("@isabella_v", "Solicité la condonación hace 6 meses y no hay respuesta", False, None, None),
                com("@andres_mz", "¿La condonación aplica solo para el crédito ACCES?", True, 520,
                    "@andres_mz aplica para ACCES y Tú Eliges. Para otros tipos de crédito hay condiciones diferentes, escríbenos al chat de atención."),
            ],
        },
    ]

    # Recalcular cumple_sla con los SLA reales de cada plataforma
    for p in fb_posts:
        for c in p["comentarios"]:
            if c["tiempo_respuesta_min"] is not None:
                c["cumple_sla"] = c["tiempo_respuesta_min"] <= 8 * 60
    for p in ig_media:
        for c in p["comentarios"]:
            if c.get("usuario", "").lstrip("@").lower() == ig_user.lower():
                continue
            if c["tiempo_respuesta_min"] is not None:
                c["cumple_sla"] = c["tiempo_respuesta_min"] <= 12 * 60

    return {"twitter": [], "facebook": fb_posts, "instagram": ig_media}


# ──────────────────────────────────────────────────────────
# MAIN
# ──────────────────────────────────────────────────────────
def main():
    modo_demo      = "--demo"        in sys.argv
    modo_desde_json = "--desde-json" in sys.argv

    print("=" * 58)
    print("   ICETEX Social Media Monitor  v2.0")
    if modo_demo:
        print("   ** MODO DEMO — datos simulados **")
    elif modo_desde_json:
        print("   ** MODO MANUAL — datos ingresados **")
    print("=" * 58)

    if not EXCEL_OK:
        print("⚠️  Excel deshabilitado. Para activarlo: pip install openpyxl\n")

    config = cargar_config()
    days  = config["settings"].get("days_back", 7)
    posts = config["settings"].get("max_posts_per_platform", 10)

    sla_config = {
        "twitter":   config.get("twitter",   {}).get("sla_horas", 4),
        "facebook":  config.get("facebook",  {}).get("sla_horas", 8),
        "instagram": config.get("instagram", {}).get("sla_horas", 12),
    }

    if modo_demo:
        fb_name = config.get("facebook", {}).get("page_id", "MiPagina")
        ig_name = config.get("instagram", {}).get("username", "mi_instagram")
        print(f"\n🎭 Cargando datos demo para Facebook='{fb_name}' / Instagram='{ig_name}'...")
        resultados = generar_datos_demo(fb_name, ig_name)
    elif modo_desde_json:
        json_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "datos_social_icetex.json")
        print(f"\n📂 Cargando datos desde: {json_path}")
        with open(json_path, encoding="utf-8") as f:
            resultados = json.load(f)
        total = sum(len(v) for v in resultados.values())
        print(f"   {total} publicaciones cargadas.")
    else:
        resultados: Dict[str, List[Dict]] = {"twitter": [], "facebook": [], "instagram": []}

        # X / Twitter
        tw = config.get("twitter", {})
        if tw.get("bearer_token", "").startswith("TU_"):
            print("\n⚠️  X: edita config.json → twitter.bearer_token")
        else:
            resultados["twitter"] = TwitterMonitor(
                tw["bearer_token"], tw.get("username", "ICETEXColombia")
            ).analizar(posts, days, sla_config["twitter"])

        # Facebook
        fb = config.get("facebook", {})
        if fb.get("access_token", "").startswith("TU_"):
            print("\n⚠️  Facebook: edita config.json → facebook.access_token")
        else:
            resultados["facebook"] = FacebookMonitor(
                fb["page_id"], fb["access_token"]
            ).analizar(posts, days, sla_config["facebook"])

        # Instagram
        ig = config.get("instagram", {})
        if not ig.get("user_id") or not ig.get("access_token"):
            print("\n⚠️  Instagram: edita config.json → instagram.user_id / access_token")
        else:
            resultados["instagram"] = InstagramMonitor(
                ig["user_id"], ig["access_token"], ig.get("username", "")
            ).analizar(posts, sla_config["instagram"])

    # Métricas
    metricas_plat = {k: calcular_metricas(v, sla_config[k]) for k, v in resultados.items() if v}
    all_pubs = [p for v in resultados.values() for p in v]
    mg = calcular_metricas(all_pubs, max(sla_config.values())) if all_pubs else {}

    base = os.path.dirname(os.path.abspath(__file__))

    # Reporte HTML
    print("\n📄 Generando reporte HTML...")
    html_out = os.path.join(base, "reporte_social_icetex.html")
    with open(html_out, "w", encoding="utf-8") as f:
        f.write(generar_reporte(resultados, config, sla_config, metricas_plat, mg))
    print(f"✅ HTML guardado en:\n   {html_out}")

    # Excel
    if EXCEL_OK:
        print("\n📊 Generando Excel...")
        xlsx_out = os.path.join(base, "reporte_social_icetex.xlsx")
        exportar_excel(resultados, metricas_plat, mg, sla_config, xlsx_out)

    # JSON respaldo (no sobreescribir si los datos vienen del JSON manual)
    if not modo_desde_json:
        json_out = os.path.join(base, "datos_social_icetex.json")
        with open(json_out, "w", encoding="utf-8") as f:
            json.dump(resultados, f, ensure_ascii=False, indent=2, default=str)
        print(f"📦 JSON en:\n   {json_out}")

    # Alertas
    print("\n📧 Verificando alertas...")
    enviar_alerta(config, resultados, html_out)

    print("\n✅ Listo.")


if __name__ == "__main__":
    main()
